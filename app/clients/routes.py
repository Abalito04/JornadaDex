import csv
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
import unicodedata

from flask import Blueprint, Response, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError

from app.context import current_company_id
from app.extensions import db
from app.models import AccountingClient
from app.permissions.decorators import manager_required
from app.services.audit_service import write_audit

clients_bp = Blueprint("clients", __name__, url_prefix="/clients")

CLIENT_IMPORT_COLUMNS = [
    ("Razon social", "name"),
    ("CUIT", "tax_id"),
    ("Direccion", "address"),
    ("Situacion IVA", "fiscal_condition"),
    ("Situacion IIBB", "multilateral_agreement"),
    ("Balance", "does_balance"),
    ("SICORE", "sicore"),
    ("Ganancias", "income_tax"),
    ("Bienes personales", "personal_assets"),
    ("Sueldos", "payroll_enabled"),
    ("Cantidad colaboradores", "payroll_employee_count"),
    ("Grupo", "group_enabled"),
    ("Nombre grupo", "group_name"),
    ("Horas presupuestadas", "budgeted_hours"),
    ("Honorarios", "fees"),
    ("Activo", "active"),
    ("Notas", "notes"),
]

HEADER_ALIASES = {
    "razon social": "name",
    "razon_social": "name",
    "nombre": "name",
    "cliente": "name",
    "cuit": "tax_id",
    "cuil": "tax_id",
    "tax_id": "tax_id",
    "direccion": "address",
    "domicilio": "address",
    "situacion iva": "fiscal_condition",
    "situacion_iva": "fiscal_condition",
    "iva": "fiscal_condition",
    "situacion iibb": "multilateral_agreement",
    "situacion_iibb": "multilateral_agreement",
    "iibb": "multilateral_agreement",
    "convenio": "multilateral_agreement",
    "balance": "does_balance",
    "se realiza balance": "does_balance",
    "sicore": "sicore",
    "ganancias": "income_tax",
    "bienes personales": "personal_assets",
    "bienes_personales": "personal_assets",
    "sueldos": "payroll_enabled",
    "cantidad colaboradores": "payroll_employee_count",
    "cantidad_colaboradores": "payroll_employee_count",
    "cantidad empleados": "payroll_employee_count",
    "cantidad_empleados": "payroll_employee_count",
    "grupo": "group_enabled",
    "nombre grupo": "group_name",
    "nombre_grupo": "group_name",
    "horas presupuestadas": "budgeted_hours",
    "horas_presupuestadas": "budgeted_hours",
    "honorarios": "fees",
    "activo": "active",
    "estado": "active",
    "notas": "notes",
    "observaciones": "notes",
}


@clients_bp.route("/")
@manager_required
def index():
    clients = (
        AccountingClient.query.filter_by(company_id=current_company_id(), deleted_at=None)
        .order_by(AccountingClient.name)
        .all()
    )
    return render_template("clients/index.html", clients=clients)


@clients_bp.route("/create", methods=["GET", "POST"])
@manager_required
def create():
    client = AccountingClient(company_id=current_company_id())
    if request.method == "POST":
        return _save_client(client, "Cliente contable creado.", "CREATE")
    return _render_form(client)


@clients_bp.route("/<int:client_id>/edit", methods=["GET", "POST"])
@manager_required
def edit(client_id):
    client = AccountingClient.query.filter_by(
        id=client_id,
        company_id=current_company_id(),
        deleted_at=None,
    ).first_or_404()
    if request.method == "POST":
        return _save_client(client, "Cliente contable actualizado.", "UPDATE")
    return _render_form(client)


@clients_bp.route("/<int:client_id>/delete", methods=["POST"])
@manager_required
def delete(client_id):
    client = AccountingClient.query.filter_by(
        id=client_id,
        company_id=current_company_id(),
        deleted_at=None,
    ).first_or_404()
    client.soft_delete(current_user.id)
    client.active = False
    write_audit("DELETE", "accounting_clients", client.id, previous_values={"name": client.name})
    db.session.commit()
    flash("Cliente contable eliminado lógicamente.", "success")
    return redirect(url_for("clients.index"))


@clients_bp.route("/import", methods=["POST"])
@manager_required
def import_clients():
    uploaded_file = request.files.get("clients_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("Selecciona un archivo CSV o XLSX para importar clientes.", "danger")
        return redirect(url_for("clients.index"))

    filename = uploaded_file.filename.lower()
    try:
        if filename.endswith(".csv"):
            rows = _read_csv_rows(uploaded_file)
        elif filename.endswith(".xlsx"):
            rows = _read_xlsx_rows(uploaded_file)
        else:
            raise ValueError("El archivo debe tener formato .csv o .xlsx.")

        created, updated, skipped, errors = _import_client_rows(rows)
        db.session.commit()
    except (IntegrityError, ValueError) as exc:
        db.session.rollback()
        flash(str(getattr(exc, "orig", exc)), "danger")
        return redirect(url_for("clients.index"))

    message = f"Importacion finalizada: {created} creados, {updated} actualizados"
    if skipped:
        message += f", {skipped} omitidos"
    flash(message + ".", "success")
    for error in errors[:5]:
        flash(error, "warning")
    if len(errors) > 5:
        flash(f"Hay {len(errors) - 5} observaciones mas no mostradas.", "warning")
    return redirect(url_for("clients.index"))


@clients_bp.route("/template.csv")
@manager_required
def template_csv():
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([header for header, _field in CLIENT_IMPORT_COLUMNS])
    content = "\ufeff" + buffer.getvalue()
    return Response(
        content,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=plantilla_clientes.csv"},
    )


@clients_bp.route("/template.xlsx")
@manager_required
def template_xlsx():
    workbook = _build_template_workbook()
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="plantilla_clientes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _save_client(client, success_message, audit_action):
    try:
        previous_values = None
        if client.id:
            previous_values = {
                "name": client.name,
                "tax_id": client.tax_id,
                "address": client.address,
                "fiscal_condition": client.fiscal_condition,
                "multilateral_agreement": client.multilateral_agreement,
                "does_balance": client.does_balance,
                "sicore": client.sicore,
                "income_tax": client.income_tax,
                "personal_assets": client.personal_assets,
                "payroll_enabled": client.payroll_enabled,
                "payroll_employee_count": client.payroll_employee_count,
                "group_enabled": client.group_enabled,
                "group_name": client.group_name,
                "budgeted_hours": str(client.budgeted_hours) if client.budgeted_hours is not None else None,
                "fees": str(client.fees) if client.fees is not None else None,
                "active": client.active,
            }

        client.name = request.form.get("name", "").strip()
        client.tax_id = request.form.get("tax_id", "").strip() or None
        client.address = request.form.get("address", "").strip() or None
        client.fiscal_condition = request.form.get("fiscal_condition", "").strip() or None
        client.multilateral_agreement = request.form.get("multilateral_agreement", "").strip() or None
        client.does_balance = _yes_no("does_balance")
        client.sicore = _yes_no("sicore")
        client.income_tax = _yes_no("income_tax")
        client.personal_assets = _yes_no("personal_assets")
        client.payroll_enabled = _yes_no("payroll_enabled")
        client.payroll_employee_count = _parse_positive_int("payroll_employee_count") if client.payroll_enabled else None
        client.group_enabled = _yes_no("group_enabled")
        client.group_name = _canonical_group_name(request.form.get("group_name", "")) if client.group_enabled else None
        client.budgeted_hours = _parse_hours("budgeted_hours")
        client.fees = _parse_currency("fees")
        client.active = request.form.get("active", "on") == "on"
        client.notes = request.form.get("notes", "").strip() or None

        if not client.name:
            raise ValueError("El nombre del cliente es obligatorio.")
        if client.payroll_enabled and client.payroll_employee_count is None:
            raise ValueError("Ingresá la cantidad de colaboradores cuando Sueldos sea Si.")
        if client.group_enabled and not client.group_name:
            raise ValueError("Ingresá el nombre de grupo cuando Grupo sea Si.")

        if not client.id:
            client.created_by = current_user.id
            db.session.add(client)
        else:
            client.updated_by = current_user.id

        db.session.flush()
        write_audit(
            audit_action,
            "accounting_clients",
            client.id,
            previous_values=previous_values,
            new_values={
                "name": client.name,
                "tax_id": client.tax_id,
                "fiscal_condition": client.fiscal_condition,
                "does_balance": client.does_balance,
                "sicore": client.sicore,
                "income_tax": client.income_tax,
                "personal_assets": client.personal_assets,
                "payroll_enabled": client.payroll_enabled,
                "payroll_employee_count": client.payroll_employee_count,
                "group_enabled": client.group_enabled,
                "group_name": client.group_name,
                "budgeted_hours": str(client.budgeted_hours) if client.budgeted_hours is not None else None,
                "fees": str(client.fees) if client.fees is not None else None,
                "active": client.active,
            },
        )
        db.session.commit()
        flash(success_message, "success")
        return redirect(url_for("clients.index"))
    except (IntegrityError, ValueError) as exc:
        db.session.rollback()
        flash(str(getattr(exc, "orig", exc)), "danger")
        return _render_form(client)


def _import_client_rows(rows):
    if rows is None:
        raise ValueError("No se pudo leer el archivo.")

    created = 0
    updated = 0
    skipped = 0
    errors = []
    company_id = current_company_id()

    for row_number, row in rows:
        if _row_is_empty(row):
            continue
        try:
            values = _client_values_from_import_row(row)
        except ValueError as exc:
            skipped += 1
            errors.append(f"Fila {row_number}: {exc}")
            continue

        client = AccountingClient.query.filter_by(company_id=company_id, name=values["name"]).first()
        previous_values = _client_snapshot(client) if client else None
        if client:
            updated += 1
            client.updated_by = current_user.id
            audit_action = "UPDATE"
        else:
            created += 1
            client = AccountingClient(company_id=company_id, created_by=current_user.id)
            db.session.add(client)
            audit_action = "CREATE"

        _apply_client_values(client, values)
        db.session.flush()
        write_audit(
            audit_action,
            "accounting_clients",
            client.id,
            previous_values=previous_values,
            new_values=_client_snapshot(client),
        )

    return created, updated, skipped, errors


def _read_csv_rows(uploaded_file):
    content = uploaded_file.stream.read().decode("utf-8-sig")
    try:
        dialect = csv.Sniffer().sniff(content[:2048], delimiters=",;")
    except csv.Error:
        dialect = csv.excel
    stream = StringIO(content)
    reader = csv.DictReader(stream, dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("El CSV no tiene encabezados.")
    normalized_headers = {header: _field_for_header(header) for header in reader.fieldnames}
    rows = []
    for index, raw_row in enumerate(reader, start=2):
        row = {}
        for header, value in raw_row.items():
            field = normalized_headers.get(header)
            if field:
                row[field] = value
        rows.append((index, row))
    return rows


def _read_xlsx_rows(uploaded_file):
    uploaded_file.stream.seek(0)
    workbook = load_workbook(uploaded_file.stream, data_only=True, read_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("El XLSX no tiene encabezados.")

    fields = [_field_for_header(header) for header in header_row]
    rows = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {}
        for field, value in zip(fields, values):
            if field:
                row[field] = value
        rows.append((index, row))
    workbook.close()
    return rows


def _field_for_header(header):
    return HEADER_ALIASES.get(_normalize_text(header))


def _normalize_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("_", " ").replace("-", " ").strip().casefold()
    return " ".join(text.split())


def _row_is_empty(row):
    return not any(_clean_cell(value) for value in row.values())


def _clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.strip().split())
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _client_values_from_import_row(row):
    name = _clean_cell(row.get("name"))
    if not name:
        raise ValueError("la razon social es obligatoria.")

    payroll_enabled = _parse_bool_value(row.get("payroll_enabled"), default=False)
    payroll_employee_count = (
        _parse_positive_int_value(row.get("payroll_employee_count"), "Cantidad colaboradores")
        if payroll_enabled
        else None
    )
    group_enabled = _parse_bool_value(row.get("group_enabled"), default=False)
    group_name = _canonical_group_name(_clean_cell(row.get("group_name"))) if group_enabled else None
    if group_enabled and not group_name:
        raise ValueError("si Grupo es Si, completa Nombre grupo.")

    return {
        "name": name,
        "tax_id": _clean_cell(row.get("tax_id")) or None,
        "address": _clean_cell(row.get("address")) or None,
        "fiscal_condition": _canonical_fiscal_condition(row.get("fiscal_condition")),
        "multilateral_agreement": _canonical_multilateral_agreement(row.get("multilateral_agreement")),
        "does_balance": _parse_bool_value(row.get("does_balance"), default=False),
        "sicore": _parse_bool_value(row.get("sicore"), default=False),
        "income_tax": _parse_bool_value(row.get("income_tax"), default=False),
        "personal_assets": _parse_bool_value(row.get("personal_assets"), default=False),
        "payroll_enabled": payroll_enabled,
        "payroll_employee_count": payroll_employee_count,
        "group_enabled": group_enabled,
        "group_name": group_name,
        "budgeted_hours": _parse_hours_value(row.get("budgeted_hours")),
        "fees": _parse_currency_value(row.get("fees")),
        "active": _parse_bool_value(row.get("active"), default=True),
        "notes": _clean_cell(row.get("notes")) or None,
    }


def _apply_client_values(client, values):
    client.name = values["name"]
    client.tax_id = values["tax_id"]
    client.address = values["address"]
    client.fiscal_condition = values["fiscal_condition"]
    client.multilateral_agreement = values["multilateral_agreement"]
    client.does_balance = values["does_balance"]
    client.sicore = values["sicore"]
    client.income_tax = values["income_tax"]
    client.personal_assets = values["personal_assets"]
    client.payroll_enabled = values["payroll_enabled"]
    client.payroll_employee_count = values["payroll_employee_count"]
    client.group_enabled = values["group_enabled"]
    client.group_name = values["group_name"]
    client.budgeted_hours = values["budgeted_hours"]
    client.fees = values["fees"]
    client.active = values["active"]
    client.notes = values["notes"]
    client.deleted_at = None
    client.deleted_by = None


def _client_snapshot(client):
    if not client:
        return None
    return {
        "name": client.name,
        "tax_id": client.tax_id,
        "address": client.address,
        "fiscal_condition": client.fiscal_condition,
        "multilateral_agreement": client.multilateral_agreement,
        "does_balance": client.does_balance,
        "sicore": client.sicore,
        "income_tax": client.income_tax,
        "personal_assets": client.personal_assets,
        "payroll_enabled": client.payroll_enabled,
        "payroll_employee_count": client.payroll_employee_count,
        "group_enabled": client.group_enabled,
        "group_name": client.group_name,
        "budgeted_hours": str(client.budgeted_hours) if client.budgeted_hours is not None else None,
        "fees": str(client.fees) if client.fees is not None else None,
        "active": client.active,
        "notes": client.notes,
    }


def _yes_no(field_name):
    return request.form.get(field_name) in {"1", "on", "true", "True", "si", "Si"}


def _render_form(client):
    return render_template("clients/form.html", client=client, group_names=_group_names())


def _group_names():
    rows = (
        db.session.query(AccountingClient.group_name)
        .filter(
            AccountingClient.company_id == current_company_id(),
            AccountingClient.deleted_at.is_(None),
            AccountingClient.group_enabled.is_(True),
            AccountingClient.group_name.isnot(None),
            AccountingClient.group_name != "",
        )
        .order_by(AccountingClient.group_name)
        .all()
    )
    names = []
    seen = set()
    for (name,) in rows:
        normalized = " ".join(name.split())
        key = normalized.casefold()
        if normalized and key not in seen:
            seen.add(key)
            names.append(normalized)
    return names


def _canonical_group_name(value):
    normalized = " ".join(value.strip().split())
    if not normalized:
        return None
    for existing_name in _group_names():
        if existing_name.casefold() == normalized.casefold():
            return existing_name
    return normalized


def _parse_positive_int(field_name):
    value = request.form.get(field_name, "").strip()
    if not value:
        return None
    try:
        parsed = int(value)
    except ValueError as exc:
        raise ValueError("La cantidad de colaboradores debe ser un número entero.") from exc
    if parsed < 0:
        raise ValueError("La cantidad de colaboradores no puede ser negativa.")
    return parsed


def _parse_decimal(field_name):
    value = request.form.get(field_name, "").strip()
    if not value:
        return None
    if "," in value and "." in value:
        value = value.replace(".", "").replace(",", ".")
    elif "," in value:
        value = value.replace(",", ".")
    try:
        parsed = Decimal(value)
    except InvalidOperation as exc:
        raise ValueError("Horas presupuestadas y honorarios deben ser valores numéricos.") from exc
    if parsed < 0:
        raise ValueError("Horas presupuestadas y honorarios no pueden ser negativos.")
    return parsed


def _parse_currency(field_name):
    value = request.form.get(field_name, "").strip().replace("$", "").replace(" ", "")
    if not value:
        return None
    if "," in value and "." in value:
        value = value.replace(".", "").replace(",", ".")
    elif "," in value:
        value = value.replace(",", ".")
    elif "." in value:
        value = value.replace(".", "")
    try:
        parsed = Decimal(value)
    except InvalidOperation as exc:
        raise ValueError("Honorarios debe ser un valor numérico.") from exc
    if parsed < 0:
        raise ValueError("Honorarios no puede ser negativo.")
    return parsed


def _parse_hours(field_name):
    value = request.form.get(field_name, "").strip().lower().removesuffix("hs").strip()
    if not value:
        return None
    if ":" in value:
        hours_text, minutes_text = value.split(":", 1)
        try:
            hours = int(hours_text or 0)
            minutes = int(minutes_text or 0)
        except ValueError as exc:
            raise ValueError("Horas presupuestadas debe tener formato 00:00.") from exc
        if hours < 0 or minutes < 0 or minutes > 59:
            raise ValueError("Horas presupuestadas debe tener formato 00:00.")
        return Decimal(hours) + (Decimal(minutes) / Decimal("60"))
    return _parse_decimal(field_name)


def _canonical_fiscal_condition(value):
    text = _clean_cell(value)
    if not text:
        return None
    options = {
        "responsable inscripto": "Responsable Inscripto",
        "monotributo": "Monotributo",
        "exento": "Exento",
        "no aplica": "No aplica",
        "no responsable": "No aplica",
    }
    return options.get(_normalize_text(text), text)


def _canonical_multilateral_agreement(value):
    text = _clean_cell(value)
    if not text:
        return None
    options = {
        "regimen general": "Regimen General",
        "regimen lateral": "Regimen General",
        "convenio multilateral": "Convenio Multilateral",
        "regimen simplificado": "Regimen Simplificado",
        "no aplica": "No Aplica",
        "no corresponde": "No Aplica",
    }
    return options.get(_normalize_text(text), text)


def _parse_bool_value(value, default=False):
    text = _normalize_text(value)
    if not text:
        return default
    if text in {"1", "si", "s", "true", "verdadero", "yes", "y", "x", "activo"}:
        return True
    if text in {"0", "no", "n", "false", "falso", "inactivo"}:
        return False
    raise ValueError(f"'{_clean_cell(value)}' debe ser Si o No.")


def _parse_positive_int(field_name):
    return _parse_positive_int_value(request.form.get(field_name, ""), "La cantidad de colaboradores")


def _parse_positive_int_value(value, label):
    value = _clean_cell(value)
    if not value:
        return None
    try:
        parsed = int(value)
    except ValueError as exc:
        raise ValueError(f"{label} debe ser un numero entero.") from exc
    if parsed < 0:
        raise ValueError(f"{label} no puede ser negativa.")
    return parsed


def _parse_decimal(field_name):
    return _parse_decimal_value(request.form.get(field_name, ""), "Horas presupuestadas y honorarios")


def _parse_decimal_value(value, label):
    if isinstance(value, (int, float, Decimal)):
        parsed = Decimal(str(value))
        if parsed < 0:
            raise ValueError(f"{label} no pueden ser negativos.")
        return parsed
    value = _clean_cell(value)
    if not value:
        return None
    if "," in value and "." in value:
        value = value.replace(".", "").replace(",", ".")
    elif "," in value:
        value = value.replace(",", ".")
    try:
        parsed = Decimal(value)
    except InvalidOperation as exc:
        raise ValueError(f"{label} deben ser valores numericos.") from exc
    if parsed < 0:
        raise ValueError(f"{label} no pueden ser negativos.")
    return parsed


def _parse_currency(field_name):
    return _parse_currency_value(request.form.get(field_name, ""))


def _parse_currency_value(value):
    if isinstance(value, (int, float, Decimal)):
        parsed = Decimal(str(value))
        if parsed < 0:
            raise ValueError("Honorarios no puede ser negativo.")
        return parsed
    value = _clean_cell(value).replace("$", "").replace(" ", "")
    if not value:
        return None
    if "," in value and "." in value:
        value = value.replace(".", "").replace(",", ".")
    elif "," in value:
        value = value.replace(",", ".")
    elif "." in value:
        value = value.replace(".", "")
    try:
        parsed = Decimal(value)
    except InvalidOperation as exc:
        raise ValueError("Honorarios debe ser un valor numerico.") from exc
    if parsed < 0:
        raise ValueError("Honorarios no puede ser negativo.")
    return parsed


def _parse_hours(field_name):
    return _parse_hours_value(request.form.get(field_name, ""))


def _parse_hours_value(value):
    value = _clean_cell(value).lower().removesuffix("hs").strip()
    if not value:
        return None
    if ":" in value:
        time_parts = value.split(":")
        hours_text = time_parts[0]
        minutes_text = time_parts[1] if len(time_parts) > 1 else "0"
        try:
            hours = int(hours_text or 0)
            minutes = int(minutes_text or 0)
        except ValueError as exc:
            raise ValueError("Horas presupuestadas debe tener formato 00:00.") from exc
        if hours < 0 or minutes < 0 or minutes > 59:
            raise ValueError("Horas presupuestadas debe tener formato 00:00.")
        return Decimal(hours) + (Decimal(minutes) / Decimal("60"))
    return _parse_decimal_value(value, "Horas presupuestadas")


def _build_template_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clientes"
    headers = [header for header, _field in CLIENT_IMPORT_COLUMNS]
    sheet.append(headers)
    sheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="0B7A53")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    widths = {
        "A": 28,
        "B": 16,
        "C": 28,
        "D": 24,
        "E": 24,
        "K": 22,
        "M": 22,
        "N": 22,
        "O": 16,
        "Q": 34,
    }
    for index in range(1, len(headers) + 1):
        column_letter = get_column_letter(index)
        sheet.column_dimensions[column_letter].width = widths.get(column_letter, 14)

    notes = workbook.create_sheet("Instrucciones")
    notes.append(["Campo", "Como completarlo"])
    for cell in notes[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    instructions = {
        "Razon social": "Obligatorio. Si ya existe en la empresa, se actualiza.",
        "CUIT": "Opcional. Puede ir con o sin guiones.",
        "Direccion": "Opcional.",
        "Situacion IVA": "Responsable Inscripto, Monotributo, Exento o No aplica.",
        "Situacion IIBB": "Regimen General, Convenio Multilateral, Regimen Simplificado o No Aplica.",
        "Balance": "Si o No.",
        "SICORE": "Si o No.",
        "Ganancias": "Si o No.",
        "Bienes personales": "Si o No.",
        "Sueldos": "Si o No.",
        "Cantidad colaboradores": "Completar solo si Sueldos es Si.",
        "Grupo": "Si o No.",
        "Nombre grupo": "Completar solo si Grupo es Si.",
        "Horas presupuestadas": "Formato 00:00, horas decimales o vacio.",
        "Honorarios": "Numero o importe, por ejemplo 150000 o $ 150.000.",
        "Activo": "Si o No. Si queda vacio, se considera Si.",
        "Notas": "Opcional.",
    }
    for header, _field in CLIENT_IMPORT_COLUMNS:
        notes.append([header, instructions.get(header, "")])
    notes.column_dimensions["A"].width = 28
    notes.column_dimensions["B"].width = 76
    return workbook
